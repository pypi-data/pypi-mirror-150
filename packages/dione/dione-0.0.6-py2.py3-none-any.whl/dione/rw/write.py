# -*- coding: utf-8 -*-
# !/usr/bin/env python
#
# Copyright 2021-2022 European Commission (JRC)
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl
"""
It contains functions and a model `dsp` to write output data into files.
"""
import io
import os
import zipfile
import logging
import os.path as osp
import schedula as sh
from . import file_ext, file_name
from .._version import __version__

log = logging.getLogger(__name__)
#: Write Model.
dsp = sh.BlueDispatcher(name='write_data', raises=True)


@sh.add_function(dsp, outputs=['results'])
def parse_outputs(outputs):
    """
    Formats the output results.

    :param outputs:
        Model outputs.
    :type outputs: schedula.utils.sol.Solution

    :return:
        Output results.
    :rtype: dict
    """
    import pandas as pd
    res = {k: v for k, v in sh.selector((
        'reduction_tech_cost', 'incompatibility_tech', 'include_exclude_tech',
        'reduction_tech_cost_pareto', 'optimal_pareto',
        'reduction_tech_cost_params', 'reduction_tech_cost_curves',
        'fleet_registrations', 'fleet_co2_references',
        'fleet_energy_references', 'fleet_targets', 'fleet_reduction_tech_cost',
        'fleet_energy', 'fleet_activity', 'fleet_energy_cost',
        'fleet_maintenance_cost', 'fleet_residual_value', 'fleet_activity_cost',
        'fleet_tech_cost', 'fleet_total_cost', 'fleet_maintenance_vat',
        'fleet_cumulative_tech_cost', 'fleet_cumulative_maintenance_cost',
        'fleet_cumulative_activity_cost', 'fleet_cumulative_total_cost',
        'fleet_margin_tech_cost', 'use_optimal_fitting', 'use_curves',
        'input_version', 'input_scope', 'input_description'
    ), outputs, allow_miss=True).items() if v is not None}
    case = res.get('include_exclude_tech')
    if case is not None:
        res['include_exclude_tech'] = case.unstack('technology')
    optimal = res.pop('optimal_pareto', None)
    if optimal is not None:
        res['reduction_tech_cost_pareto'] = pd.concat([
            res['reduction_tech_cost_pareto'], optimal
        ], axis=1)

    tech = res.get('incompatibility_tech')
    if tech is not None and not tech.empty:
        tech = pd.Series({
            (k, i): 'x' for k, v in tech.items() for i in v
        }).unstack()
        tech.index.name = 'technology'
        res['incompatibility_tech'] = tech
    groups = {
        'fleet_residual_maintenance': (
            'fleet_maintenance_cost', 'fleet_residual_value',
            'fleet_margin_tech_cost', 'fleet_maintenance_vat'
        ),
        'fleet_tco': (
            'fleet_activity_cost', 'fleet_tech_cost', 'fleet_total_cost',
            'fleet_cumulative_maintenance_cost', 'fleet_cumulative_tech_cost',
            'fleet_cumulative_activity_cost', 'fleet_cumulative_total_cost',
        )
    }
    inputs = {
        k: res.pop(k)
        for k, v in tuple(res.items())
        if not isinstance(v, (pd.Series, pd.DataFrame))
    }
    inputs['version'] = __version__
    for k, v in groups.items():
        v = [res.pop(i) for i in v if i in res]
        if v:
            names = v[0].index.names
            v = [v[0]] + [d.reorder_levels(names) for d in v[1:]]
            res[k] = pd.concat(v, axis=1).dropna()
    res['inputs'] = inputs
    return res


@sh.add_function(dsp, outputs=['output_fpath'])
def default_output_fpath(out_dir):
    import datetime
    return osp.join(
        out_dir, datetime.datetime.today().strftime('%Y%m%d_%H%M%S-output.xlsx')
    )


@sh.add_function(dsp, input_domain=file_ext('xlsx'), outputs=['written'])
def save_excel_template(template_fpath):
    """
    Save template as excel.

    :param template_fpath:
        Template file path.
    :type template_fpath: str
    """
    import shutil
    os.makedirs(osp.dirname(template_fpath) or '.', exist_ok=True)
    shutil.copy(osp.join(
        osp.dirname(osp.dirname(__file__)), 'templates', 'input_template.xlsx'
    ), template_fpath)
    return template_fpath


@sh.add_function(dsp, outputs=['written'])
def save_excel_demos(demo_fpath):
    """
    Save demo as excel.

    :param demo_fpath:
        Demo file path.
    :type demo_fpath: str
    """
    log.info(f'Saving demo file {demo_fpath}...')
    import shutil
    os.makedirs(osp.dirname(demo_fpath) or '.', exist_ok=True)
    shutil.copy(
        osp.join(osp.dirname(osp.dirname(__file__)), 'demos', 'demo.xlsx'),
        demo_fpath
    )
    log.info(f'Saved demo file {demo_fpath}!')
    return demo_fpath


# noinspection PyUnusedLocal
@sh.add_function(dsp, input_domain=file_ext('bz2'), outputs=['output_file'])
def save_zip(output_fpath, results):
    """
    Save output results in an Excel file.

    :param output_fpath:
        Output file path.
    :type output_fpath: str

    :param results:
        Output results.
    :type results: pandas.DataFrame

    :return:
        Output data as BinaryIO.
    :rtype: io.BytesIO
    """
    import pyarrow.feather as feather

    log.info(f'Saving bz2 file {osp.basename(output_fpath)}...')
    fd = io.BytesIO()

    with zipfile.ZipFile(fd, mode='w', compression=zipfile.ZIP_BZIP2) as z:
        sheets = {
            'reduction_tech_cost': [
                'technology', 'powertrain', 'segment', 'registration_year',
                'cost_type', 'reduction', 'cost',
            ],
            'incompatibility_tech': ['technology'],
            'include_exclude_tech': [
                'powertrain', 'segment', 'registration_year', 'cost_type',
                'technology'
            ],
            'reduction_tech_cost_pareto': [
                'powertrain', 'segment', 'registration_year', 'cost_type',
                'reduction', 'cost', 'optimal', 'technologies'
            ],
            'reduction_tech_cost_params': [
                'powertrain', 'segment', 'registration_year', 'cost_type', 'A',
                'B', 'C', 'c', 'x0', 'BA', 'BSC', 'TO', 'RCO2', 'RC', 'x_max',
                'x_min', 'y_max', 'y_min'
            ],
            'reduction_tech_cost_curves': [
                'powertrain', 'segment', 'registration_year', 'cost_type', 'A',
                'B', 'C', 'c', 'x0', 'x_max', 'x_min', 'y_max', 'y_min', 'MAE'
            ],
            'fleet_registrations': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'registration_year'
            ],
            'fleet_co2_references': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type', 'co2'
            ],
            'fleet_targets': [
                'scenario', 'fleet', 'cost_type', 'registration_year', 'co2'
            ],
            'fleet_reduction_tech_cost': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'registration_year', 'reduction', 'cost', 'shares', 'valid'
            ],
            'fleet_energy_references': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'registration_year', 'projection_year', 'conventional',
                'electric'
            ],
            'fleet_energy': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'registration_year', 'projection_year', 'conventional',
                'electric'
            ],
            'fleet_activity': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'perspective', 'registration_year', 'projection_year',
                'conventional', 'electric'
            ],
            'fleet_energy_cost': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'perspective', 'registration_year', 'projection_year',
                'conventional', 'electric'
            ],
            'fleet_residual_maintenance': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'perspective', 'registration_year', 'age', 'residual_value',
                'margin_tech_cost', 'maintenance_cost', 'maintenance_vat'
            ],
            'fleet_tco': [
                'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
                'perspective', 'registration_year', 'age', 'tech_cost',
                'activity_cost', 'total_cost', 'cumulative_tech_cost',
                'cumulative_activity_cost', 'cumulative_maintenance_cost',
                'cumulative_total_cost'
            ]
        }
        for k, columns in sheets.items():
            if k not in results:
                continue
            df = results[k].sort_index().reset_index()
            for i in columns:
                if i not in df:
                    df[i] = ''
            if k == 'reduction_tech_cost_pareto':
                df.sort_values(columns, inplace=True)
            columns += sorted(set(df.columns) - set(columns))
            with z.open(k, mode='w') as f:
                feather.write_feather(df[columns], f)
        if 'inputs' in results:
            import json
            with z.open(k, mode='wb') as f:
                f.write(json.dumps(
                    results['inputs'], default=_json_default,
                    separators=(',', ':')
                ).encode())
    log.info(f'Saved bz2 file {osp.basename(output_fpath)}!')
    return fd


# noinspection PyUnusedLocal
@sh.add_function(dsp, input_domain=file_ext('xlsx'), outputs=['output_file'])
def save_excel(output_fpath, results):
    """
    Save output results in an Excel file.

    :param output_fpath:
        Output file path.
    :type output_fpath: str

    :param results:
        Output results.
    :type results: pandas.DataFrame

    :return:
        Output data as BinaryIO.
    :rtype: io.BytesIO
    """
    log.info(f'Saving excel file {osp.basename(output_fpath)}...')
    fd = io.BytesIO()
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.load_workbook(osp.join(
        osp.dirname(osp.dirname(__file__)), 'templates', 'input_template.xlsx'
    ))

    sheets = {
        'reduction_tech_cost': [
            'technology', 'powertrain', 'segment', 'registration_year',
            'cost_type', 'reduction', 'cost',
        ],
        'incompatibility_tech': ['technology'],
        'include_exclude_tech': [
            'powertrain', 'segment', 'registration_year', 'cost_type',
            'technology'
        ],
        'reduction_tech_cost_pareto': [
            'powertrain', 'segment', 'registration_year', 'cost_type',
            'reduction', 'cost', 'optimal', 'technologies'
        ],
        'reduction_tech_cost_params': [
            'powertrain', 'segment', 'registration_year', 'cost_type', 'A',
            'B', 'C', 'c', 'x0', 'BA', 'BSC', 'TO', 'RCO2', 'RC', 'x_max',
            'x_min', 'y_max', 'y_min'
        ],
        'reduction_tech_cost_curves': [
            'powertrain', 'segment', 'registration_year', 'cost_type', 'A',
            'B', 'C', 'c', 'x0', 'x_max', 'x_min', 'y_max', 'y_min', 'MAE'
        ],
        'fleet_registrations': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'registration_year'
        ],
        'fleet_co2_references': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type', 'co2'
        ],
        'fleet_targets': [
            'scenario', 'fleet', 'cost_type', 'registration_year', 'co2'
        ],
        'fleet_reduction_tech_cost': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'registration_year', 'reduction', 'cost', 'shares', 'valid'
        ],
        'fleet_energy_references': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'registration_year', 'projection_year', 'conventional',
            'electric'
        ],
        'fleet_energy': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'registration_year', 'projection_year', 'conventional',
            'electric'
        ],
        'fleet_activity': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'perspective', 'registration_year', 'projection_year',
            'conventional', 'electric'
        ],
        'fleet_energy_cost': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'perspective', 'registration_year', 'projection_year',
            'conventional', 'electric'
        ],
        'fleet_residual_maintenance': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'perspective', 'registration_year', 'age', 'residual_value',
            'margin_tech_cost', 'maintenance_cost', 'maintenance_vat'
        ],
        'fleet_tco': [
            'scenario', 'fleet', 'powertrain', 'segment', 'cost_type',
            'perspective', 'registration_year', 'age', 'tech_cost',
            'activity_cost', 'total_cost', 'cumulative_tech_cost',
            'cumulative_activity_cost', 'cumulative_maintenance_cost',
            'cumulative_total_cost'
        ]
    }
    for k, columns in sheets.items():
        if k not in results:
            continue
        ws = wb[k]
        df = results[k].sort_index().reset_index()

        for i in columns:
            if i not in df:
                df[i] = ''
        if k == 'reduction_tech_cost_pareto':
            df.sort_values(columns, inplace=True)
        columns += sorted(set(df.columns) - set(columns))

        for r, row in enumerate(dataframe_to_rows(df[columns], index=False), 3):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c, value=value)

    if 'inputs' in results:
        sn = wb.sheetnames[list(map(str.lower, wb.sheetnames)).index('inputs')]
        ws = wb[sn]
        r = 2
        inputs = dict(results['inputs'])
        name = ''
        while True:
            name = ws.cell(row=r, column=2).value
            if name is None:
                break
            if name in inputs:
                ws.cell(row=r, column=3, value=inputs.pop(name))
            r += 1
        for r, (k, v) in enumerate(sorted(inputs.items()), r):
            ws.cell(row=r, column=2, value=k)
            ws.cell(row=r, column=3, value=v)

    wb.save(fd)
    log.info(f'Saved excel file {osp.basename(output_fpath)}!')
    return fd


def _json_default(o):
    import numpy as np
    import pandas as pd
    if isinstance(o, np.ndarray):
        return o.tolist()
    elif isinstance(o, pd.DataFrame):
        return o.to_dict(orient='split')
    elif isinstance(o, pd.Series):
        return o.to_frame(name='value').to_dict(orient='split')


# noinspection PyUnusedLocal
@sh.add_function(dsp, input_domain=file_ext('json'), outputs=['output_file'])
def save_json(output_fpath, results):
    """
    Save output results in an JSON file.

    :param output_fpath:
        Output file path.
    :type output_fpath: str

    :param results:
        Output results.
    :type results: pandas.DataFrame

    :return:
        Output data as BinaryIO.
    :rtype: io.BytesIO
    """
    import json
    fd = io.BytesIO()
    fd.write(json.dumps(
        results, default=_json_default, separators=(',', ':')
    ).encode())
    return fd


@sh.add_function(dsp, input_domain=file_name, outputs=['written'])
def write_file(output_fpath, output_file):
    """
    Save output results in a file.

    :param output_fpath:
        Output file path.
    :type output_fpath: str

    :param output_file:
        Output data as BinaryIO.
    :type output_file: io.BytesIO

    :return:
        File path where output are written.
    :rtype: str
    """
    os.makedirs(osp.dirname(output_fpath) or '.', exist_ok=True)
    output_file.seek(0)
    with open(output_fpath, 'wb') as f:
        f.write(output_file.read())
    return output_fpath
